"""
Tests for the PDF export Turkish character fix.
Verifies:
 - /api/ health
 - /api/export/pdf renders Turkish chars correctly (bundled DejaVu font)
 - /api/export/pdf with filters still produces a valid PDF
 - /api/export/excel non-regression (3 sheets, valid xlsx)
"""
import io
import os
import pytest
import requests
from pypdf import PdfReader
from openpyxl import load_workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/") if os.environ.get("REACT_APP_BACKEND_URL") else "https://traffic-kasko-admin.preview.emergentagent.com"
ADMIN_EMAIL = "admin@sigorta.com"
ADMIN_PASSWORD = "admin123"

TURKISH_LOWER = ["ı", "ç", "ş", "ğ", "ü", "ö"]
TURKISH_UPPER = ["İ", "Ç", "Ş", "Ğ", "Ü", "Ö"]
FALLBACK_GLYPHS = ["■", "□"]


@pytest.fixture(scope="module")
def token():
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=30,
    )
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    data = r.json()
    tok = data.get("token") or data.get("access_token")
    assert tok, f"no token in response: {data}"
    return tok


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


# --- health -----------------------------------------------------------------
def test_api_root_health():
    r = requests.get(f"{BASE_URL}/api/", timeout=30)
    assert r.status_code == 200, f"/api/ health failed: {r.status_code} {r.text[:200]}"


# --- PDF export: Turkish characters ----------------------------------------
def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    parts = []
    for p in reader.pages:
        try:
            parts.append(p.extract_text() or "")
        except Exception as e:
            parts.append("")
    return "\n".join(parts)


def test_export_pdf_returns_valid_pdf(auth_headers):
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, timeout=120)
    assert r.status_code == 200, f"pdf export failed: {r.status_code} {r.text[:300]}"
    ct = r.headers.get("content-type", "")
    assert "pdf" in ct.lower(), f"unexpected content-type: {ct}"
    assert r.content.startswith(b"%PDF"), "response is not a PDF (bad magic bytes)"
    assert len(r.content) > 2000, f"pdf too small ({len(r.content)} bytes) - likely empty"

    # persist for debugging
    with open("/tmp/export_all.pdf", "wb") as f:
        f.write(r.content)


def test_export_pdf_contains_turkish_characters(auth_headers):
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, timeout=120)
    assert r.status_code == 200
    text = _extract_pdf_text(r.content)
    assert text.strip(), "pdf text extraction returned empty"

    # Section headings written in code always have Turkish chars
    assert "Kayıt Listesi" in text, f"'Kayıt Listesi' missing – Turkish 'ı' not rendered. sample={text[:500]!r}"
    assert "Oluşturulma" in text, "'Oluşturulma' missing – 'ş' not rendered"
    assert "Aylık Onay Adetleri" in text, "'Aylık Onay Adetleri' missing"
    assert "Danışman Performansı" in text, "'Danışman Performansı' missing"
    assert "Marka Bazlı Toplamlar" in text, "'Marka Bazlı Toplamlar' missing"
    assert "Detaylı Kayıt Listesi" in text, "'Detaylı Kayıt Listesi' missing"

    # column headers with Turkish
    assert "Müşteri" in text, "'Müşteri' column header missing"
    assert "ŞASİ" in text, "'ŞASİ' column header missing"
    assert "TRAFİK" in text, "'TRAFİK' column header missing"
    assert "Danışman" in text, "'Danışman' column header missing"


def test_export_pdf_lowercase_and_uppercase_turkish_glyphs(auth_headers):
    """For each of the 6 Turkish letters, at least the lower- OR upper-case
    variant must appear in the extracted PDF text. If the DejaVu font is
    loaded, both cases render correctly; a missing letter here would mean
    the glyph fell back to a placeholder."""
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, timeout=120)
    assert r.status_code == 200
    text = _extract_pdf_text(r.content)

    pairs = list(zip(TURKISH_LOWER, TURKISH_UPPER))
    for low, up in pairs:
        assert (low in text) or (up in text), (
            f"Neither lowercase {low!r} nor uppercase {up!r} present in PDF – "
            f"font likely did not render this Turkish letter."
        )

    # sanity: uppercase İ, Ş, Ç, Ğ, Ü, Ö all present in real data / headers
    for ch in ["İ", "Ş", "Ç", "Ğ", "Ü", "Ö"]:
        assert ch in text, f"uppercase Turkish char {ch!r} missing from PDF"

    # sanity: lowercase ones that are used in code-level strings
    # ("Aylık", "Oluşturulma", "Müşteri", "Danışman")
    for ch in ["ı", "ş", "ü"]:
        assert ch in text, f"lowercase Turkish char {ch!r} missing from PDF"


def test_export_pdf_no_fallback_replacement_glyphs(auth_headers):
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, timeout=120)
    assert r.status_code == 200
    text = _extract_pdf_text(r.content)
    for g in FALLBACK_GLYPHS:
        assert g not in text, (
            f"Fallback glyph {g!r} present in PDF text – font fallback still happening"
        )


# --- PDF export: filters ----------------------------------------------------
@pytest.mark.parametrize("params", [
    {"durum": "ONAY"},
    {"durum": "RED"},
    {"year": 2026},
    {"year": 2026, "month": 7},
])
def test_export_pdf_with_filters(auth_headers, params):
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, params=params, timeout=120)
    assert r.status_code == 200, f"filter {params} failed: {r.status_code} {r.text[:200]}"
    assert r.content.startswith(b"%PDF"), f"filter {params} did not return a PDF"
    assert len(r.content) > 1500, f"filter {params} pdf too small: {len(r.content)} bytes"

    # sanity: still contains header text with Turkish chars
    text = _extract_pdf_text(r.content)
    assert "Kayıt Listesi" in text, f"'Kayıt Listesi' missing for filter {params}"
    for g in FALLBACK_GLYPHS:
        assert g not in text, f"fallback glyph {g!r} in filtered pdf {params}"


def test_export_pdf_with_marka_and_danisman_filter(auth_headers):
    # first read filters to get a real value
    f = requests.get(f"{BASE_URL}/api/filters", headers=auth_headers, timeout=30)
    assert f.status_code == 200, f
    filters = f.json()
    markalar = filters.get("markalar") or []
    danismanlar = filters.get("danismanlar") or []
    if not markalar or not danismanlar:
        pytest.skip("no markalar/danismanlar in DB to filter on")
    params = {"marka": markalar[0], "danisman": danismanlar[0]}
    r = requests.get(f"{BASE_URL}/api/export/pdf", headers=auth_headers, params=params, timeout=120)
    assert r.status_code == 200, f"combined filter failed: {r.status_code} {r.text[:200]}"
    assert r.content.startswith(b"%PDF")
    text = _extract_pdf_text(r.content)
    for g in FALLBACK_GLYPHS:
        assert g not in text, f"fallback glyph {g!r} in combined-filter pdf"


# --- Excel export: non-regression ------------------------------------------
def test_export_excel_returns_xlsx_with_three_sheets(auth_headers):
    r = requests.get(f"{BASE_URL}/api/export/excel", headers=auth_headers, timeout=120)
    assert r.status_code == 200, f"excel export failed: {r.status_code} {r.text[:200]}"
    ct = r.headers.get("content-type", "").lower()
    assert "spreadsheetml" in ct or "excel" in ct or "octet-stream" in ct, f"bad content-type {ct}"
    # xlsx magic bytes = PK zip
    assert r.content[:2] == b"PK", "not a valid xlsx (missing PK header)"

    wb = load_workbook(io.BytesIO(r.content))
    names = wb.sheetnames
    for expected in ["Kayitlar", "Danisman Performansi", "Marka Toplamlari"]:
        assert expected in names, f"missing sheet {expected!r} in workbook (got {names})"
